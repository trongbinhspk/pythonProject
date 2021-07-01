import smtplib
import mimetypes
from email.mime.multipart import MIMEMultipart
from email import encoders
from email.message import Message
from email.mime.audio import MIMEAudio
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.text import MIMEText

from openpyxl import load_workbook


### Bật Quyền truy cập của ứng dụng kém an toàn
### https://stackoverflow.com/questions/31018650/send-mail-with-python-using-gmail-smtp

### Cấu hình
emailfrom = "trongbinhspk@gmail.com"
password = "binhtrongtran"
# Kết nối gmail
server = smtplib.SMTP("smtp.gmail.com:587")
server.starttls()
server.login(emailfrom,password)
## Đọc file xlsx
wb = load_workbook(filename = 'Mail_Content.xlsx')
sheet = wb.active
max_row = sheet.max_row
for r in range(2,max_row+1):
    ### Soạn nội dung
    msg = MIMEMultipart()
    msg["From"] = emailfrom
    emailto = sheet.cell(row=r,column=1).value
    msg["To"] = emailto #emailto
    msg["Subject"] = sheet.cell(row=r,column=2).value #html
    html = sheet.cell(row=r, column=3).value  # html
    content = MIMEText(html, 'html')
    msg.attach(content)
### Đính kèm file vào nội dung
    if sheet.cell(row=r, column=4).value is not None:
        fileToSend = sheet.cell(row=r, column=4).value
        ctype, encoding = mimetypes.guess_type(fileToSend)
        if ctype is None or encoding is not None:
            ctype = "application/octet-stream"
        maintype, subtype = ctype.split("/", 1)
        if maintype == "text":
            fp = open(fileToSend)
            # Note: we should handle calculating the charset
            attachment = MIMEText(fp.read(), _subtype=subtype)
            fp.close()
        elif maintype == "image":
            fp = open(fileToSend, "rb")
            attachment = MIMEImage(fp.read(), _subtype=subtype)
            fp.close()
        elif maintype == "audio":
            fp = open(fileToSend, "rb")
            attachment = MIMEAudio(fp.read(), _subtype=subtype)
            fp.close()
        else:
            fp = open(fileToSend, "rb")
            attachment = MIMEBase(maintype, subtype)
            attachment.set_payload(fp.read())
            fp.close()
            encoders.encode_base64(attachment)
        attachment.add_header("Content-Disposition", "attachment", filename=fileToSend)
        msg.attach(attachment)
    server.sendmail(emailfrom, emailto, msg.as_string())
## nghỉ xíu
server.quit()


