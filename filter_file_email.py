import openpyxl
import pandas as pd
# read by default 1st sheet of an excel file
from openpyxl import load_workbook
wb = load_workbook(filename = 'Book1.xlsx')
sheet = wb.active
max_row = sheet.max_row

column_filter = 9 # Cột giá trị muốn lọc
print(max_row)
for r in range(1, max_row+1):
    filter_value = sheet.cell(row=r, column=1).value
    print(filter_value)
    df = pd.read_excel('filter_file.xlsx',usecols='A:J',)
    print()
    dataframe_filter = df[df[df.columns[column_filter]]==filter_value]
    #print(dataframe_filter)
    chuoi = ""
    for i in range(0, len(dataframe_filter)):
        chuoi +="<th>"
        for j in range(0, len(dataframe_filter.columns)):
            chuoi +=  "<td>"+str(dataframe_filter.iloc[i, j])+ "</td>"
        chuoi +="</th>"
    print(chuoi)




