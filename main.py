import os
import openpyxl
import pandas as pd
import xlrd



dir_containing_files = r'C:\Users\trong\Desktop\tiêm văc xin\2-DS đơn vị'


filename1 =r'C:\Users\trong\Desktop\test.xlsx'
wb2 = openpyxl.load_workbook(filename1)
ws2 = wb2.active

for root, dir, filenames in os.walk(dir_containing_files):
    for file in filenames:
        if file.endswith('.xlsx'):
            file_name = file.split('.')[0]
            print(os.path.abspath(os.path.join(root, file)))
            wb = xlrd.open_workbook(os.path.abspath(os.path.join(root, file)))
            wb1 = openpyxl.load_workbook()
            ws1 = wb1.active
            i = 10
            j = 1
            print(len(ws2['A']))

            while i<len(ws2['A']):
                    for x in range(1, 18):
                        c = ws1.cell(row=i, column=x)
                        ws2.cell(row=len(ws2['A'])+1, column=x).value = c.value
                    i=i+1
                    print("dang chy")

wb2.save(str(filename1))


